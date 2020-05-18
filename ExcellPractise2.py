from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

workbook = load_workbook(filename="src/hello_world.xlsx")
# print(workbook.sheetnames)
sheet = workbook.active
# sheet
# print(sheet)
# print(sheet.title)
# print(sheet["A1"])
# print(sheet["A1"].value)
sheet["C1"]="Lol kek cheburek!"


cell1 = "A2"
cell2 = "A80"
cell3 = "A2"
cell4 = "AG"
text = "SomoHomo"
cell = cell3
# af.getColumnNumberDecimal(cell2)
# for i in range(af.getRowNumber(cell1), af.getRowNumber(cell2) + 1):
#     for j in range(af.getColumnNumberDecimal(cell3), af.getColumnNumberDecimal(cell4)+ 1):
#         sheet[cell] = text
#         cell = af.incColumnNumber(cell)
#
#     cell1 = af.incRowNumber(cell1)
#     cell = cell1
#         # print(cell1)
#
# workbook.save(filename="hello_world.xlsx")



def readFromTableCell(filePath: str, tableCellNumber: str):
    workbook = load_workbook(filename=filePath)
    sheet = workbook.active
    return sheet[tableCellNumber].value



def writeToTableCell(filePath: str, tableCellNumber: str, valueToWrite: str):
    workbook = load_workbook(filename=filePath)
    sheet = workbook.active
    sheet[tableCellNumber] = valueToWrite
    return



from src.ExcelManager import *
from src.TableCellHelper import *

print(TableCellHelper.columnNumberToDecimal("A"))

try:
    f = ExcelManager("src/hello_world.xlsx")
    print(f.readFromTableCell("A1"))
    f.writeToTableCell("A2", "Bjorn!")
    f.save()
except InvalidFileException:
    print("Exception")
